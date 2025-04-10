import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from openpyxl import Workbook, load_workbook
import os

from playwright_bot import values

product_info = {}

def create_or_load_excel(template_path):
    """Function to create a new file or load an existing one"""
    if os.path.exists(template_path):
        wb = load_workbook(template_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Color", "Memory", "Seller", "Price", "Discount Price", "Images", "Product Code", "Reviews", "Characteristics"])
        wb.save(template_path)
    return wb, ws

def save_to_excel(product_info):
    """Function to save the data to Excel"""
    template_path = "templates/selenium_template.xlsx"
    wb, ws = create_or_load_excel(template_path)

    # Add a new row with the data
    ws.append([
        product_info.get("name"),
        product_info.get("color"),
        product_info.get("memory"),
        product_info.get("seller"),
        product_info.get("price"),
        product_info.get("discount_price"),
        ', '.join(product_info.get("image_urls", [])),
        product_info.get("product_code"),
        product_info.get("reviews"),
        ', '.join([f"{key}: {value}" for key, value in product_info.get("characteristics", {}).items()])
    ])

    # Save the changes
    wb.save(template_path)

# Configure Chrome options
chrome_options = Options()
# chrome_options.add_argument("--headless")  # Run in headless mode (no GUI)
chrome_options.add_argument("--disable-notifications")
chrome_options.add_experimental_option('detach', True)

# Set the path to your ChromeDriver
DRIVER_PATH = "../Playwright_bot/chromedriver.exe"

# Create the WebDriver service
service = Service(executable_path=DRIVER_PATH)
driver = webdriver.Chrome(service=service, options=chrome_options)

# driver.maximize_window()
driver.get("https://rozetka.com.ua/")
time.sleep(2)

# Find the search input and search for the product
search_input = driver.find_element(By.NAME, "search")
search_input.send_keys("Apple iPhone 15 128GB Black")
search_input.send_keys(Keys.RETURN)

time.sleep(2)

# Try to click on the first product link
try:
    driver.find_element(By.XPATH, "//rz-indexed-link[@class='product-link']").click()
except NoSuchElementException:
    print("Error: Product not found")
    driver.quit()
    exit()

time.sleep(2)

# Get data from the product page
try:
    product_info["name"] = driver.find_element(By.XPATH, "//h1").text
except NoSuchElementException:
    product_info["name"] = None

try:
    color_text = driver.find_element(By.XPATH, "//p[@class='text-base mb-2']").text
    product_info["color"] = color_text.replace("Колір:", "").strip() if color_text else None
except NoSuchElementException:
    product_info["color"] = None

try:
    product_info["memory"] = driver.find_element(By.XPATH, "//rz-var-parameters/rz-var-parameter-option[2]/div/p/span[@class='bold']").text
except NoSuchElementException:
    product_info["memory"] = None

try:
    product_info["seller"] = driver.find_element(By.XPATH, "//p[@class='seller-title']/span[@class='seller-logo']/img").get_attribute("alt")
except AttributeError:
    product_info["seller"] = None

try:
    product_info["price"] = driver.find_element(By.XPATH, "//p[@class='product-price__big product-price__big-color-red']").text
except NoSuchElementException:
    product_info["price"] = None

try:
    product_info["discount_price"] = driver.find_element(By.XPATH, "//p[@class='product-price__small']").text
except NoSuchElementException:
    product_info["discount_price"] = None

try:
    images = driver.find_elements(By.XPATH, "//img[@class='image']")
    product_info["image_urls"] = [img.get_attribute("src") for img in images]
except NoSuchElementException:
    product_info["image_urls"] = None

time.sleep(2)
try:
    product_text_code = driver.find_elements(By.XPATH, "//span[@class='ms-auto color-black-60']")
    product_info["product_code"] = ''.join([char.text for char in product_text_code]).replace("Код: ", "").strip()
except NoSuchElementException:
    product_info["product_code"] = None
#

try:
    product_info["reviews"] = driver.find_element(By.XPATH, "//li[@class='tabs__item'][3]/rz-indexed-link/a/span").text
except NoSuchElementException:
    product_info["reviews"] = None

try:
    characteristics_elements = driver.find_elements(By.XPATH, "//dl/div[@class='item']/dd/ul/li/rz-indexed-link/a/span")
    product_info["characteristics"] = [char.text for char in characteristics_elements]
except NoSuchElementException:
    product_info["characteristics"] = None

# navigation to the characteristics page
driver.find_element(By.XPATH, "//rz-indexed-link[@class='tabs__link']/a").click()
time.sleep(2)

try:
    keys = driver.find_elements(By.XPATH, "//dt[@class='label']")
    keys_array = [key.text for key in keys]
except NoSuchElementException:
    keys_array = None

try:
    values = driver.find_elements(By.XPATH, "//dd[@class='value']")
    values_array = [value.text for value in values]
except NoSuchElementException:
    values_array = None

if keys_array and values_array:
    product_info['characteristics'] = dict(zip(keys_array, values_array))
else:
    product_info['characteristics'] = {}


driver.quit()

for key, value in product_info.items():
    print(f"{key}: {value}")

# Save the results to Excel
save_to_excel(product_info)

print("Data successfully saved to 'templates/selenium_template.xlsx'.")
# print(new_product_code)
