import time
from playwright.sync_api import sync_playwright, TimeoutError
from openpyxl import Workbook, load_workbook
import os

product_info = {}

def create_or_load_excel(template_path):
    """Function to create a new file or load an existing one"""
    if os.path.exists(template_path):
        wb = load_workbook(template_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Color", "Memory", "Seller", "Price", "Discount Price", "Images", "Product Code", "Reviews", "Characteristics"])
        wb.save(template_path)
    return wb, ws

def save_to_excel(product_info):
    """Function to save the data to Excel"""
    template_path = "templates/playwright_template.xlsx"
    wb, ws = create_or_load_excel(template_path)

    # Add a new row with the data
    ws.append([
        product_info.get("name"),
        product_info.get("color"),
        product_info.get("memory"),
        product_info.get("seller"),
        product_info.get("price"),
        product_info.get("discount_price"),
        ', '.join(product_info.get("image_urls", [])),
        product_info.get("product_code"),
        product_info.get("reviews"),
        ', '.join([f"{key}: {value}" for key, value in product_info.get("characteristics", {}).items()])
    ])

    # Save the changes
    wb.save(template_path)

with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    page = browser.new_page()

    page.goto("https://rozetka.com.ua/")
    page.wait_for_timeout(1000)

    page.fill("input[name='search']", "Apple iPhone 15 128GB Black")

    page.locator("//button[contains(text(),'Знайти')]").click()

    try:
        page.locator("//rz-indexed-link[@class='product-link']").first.click()
    except TimeoutError:
        print("Error: Product not found")
        browser.close()
        exit()

    page.wait_for_timeout(2000)

    # Get data
    try:
        product_info["name"] = page.locator("//h1").text_content()
    except (TimeoutError, AttributeError):
        product_info["name"] = None

    try:
        color_text = page.locator("//p[@class='text-base mb-2']").first.text_content()
        product_info["color"] = color_text.replace("Колір:", "").strip() if color_text else None
    except (TimeoutError, AttributeError):
        product_info["color"] = None

    try:
        product_info["memory"] = page.locator(
            "//rz-var-parameters/rz-var-parameter-option[2]/div/p/span[@class='bold']"
        ).text_content()
    except (TimeoutError, AttributeError):
        product_info["memory"] = None

    try:
        product_info["seller"] = page.locator("//p[@class='seller-title']/span[@class='seller-logo']/img").get_attribute("alt")
    except (TimeoutError, AttributeError):
        product_info["seller"] = None

    try:
        product_info["price"] = page.locator("//p[@class='product-price__big product-price__big-color-red']").text_content()
    except (TimeoutError, AttributeError):
        product_info["price"] = None

    try:
        product_info["discount_price"] = page.locator("//p[@class='product-price__small']").text_content()
    except (TimeoutError, AttributeError):
        product_info["discount_price"] = None

    try:
        images = page.locator("//img[@class='image']").all()
        product_info["image_urls"] = [img.get_attribute("src") for img in images]
    except (TimeoutError, AttributeError):
        product_info["image_urls"] = []

    try:
        product_info["product_code"] = page.locator("//div[@class='rating text-base']/span").first.text_content()
        product_info["product_code"] = product_info["product_code"].replace("Код: ", "").strip()
    except (TimeoutError, AttributeError):
        product_info["product_code"] = None

    try:
        product_info["reviews"] = page.locator("//li[@class='tabs__item'][3]/rz-indexed-link/a/span").text_content()
    except (TimeoutError, AttributeError):
        product_info["reviews"] = None


    # navigation to the characteristics page
    time.sleep(2)
    page.locator("//rz-indexed-link[@class='tabs__link']/a").first.click()
    time.sleep(2)


    try:
        keys = page.locator("//dt[@class='label']").all()
        keys_array = [key.text_content() for key in keys]
    except AttributeError as e:
        keys_array = []
        print(f"Error extracting keys: {e}")

    try:
        values = page.locator("//dd[@class='value']").all()
        values_array = [value.text_content() for value in values]
    except AttributeError as e:
        values_array = []
        print(f"Error extracting values: {e}")

    if keys_array and values_array:
        product_info['characteristics'] = dict(zip(keys_array, values_array))
    else:
        product_info['characteristics'] = {}

    browser.close()

for key, value in product_info.items():
    print(f"{key}: {value}")

# Save results to Excel
save_to_excel(product_info)

print("Data successfully saved to 'templates/playwright_template.xlsx'.")